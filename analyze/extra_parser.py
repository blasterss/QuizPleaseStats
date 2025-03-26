import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# url с результатами игр по классике с форматирующимся номером страницы
base_url = 'https://spb.quizplease.ru/schedule-past?QpGameSearch%5BcityId%5D=17&QpGameSearch%5Bmonth%5D=0&QpGameSearch%5Btype%5D=6&QpGameSearch%5Bbars%5D=all'

winter_final_url = "https://spb.quizplease.ru/game-page?id=73339"
spring_final_url = "https://spb.quizplease.ru/game-page?id=79432"
summer_final_url = "https://spb.quizplease.ru/game-page?id=85005"


wb = Workbook()
ws = wb.active
months = {
    'января,': '01',
    'февраля,': '02',
    'марта,': '03',
    'апреля,': '04',
    'мая,': '05',
    'июня,': '06',
    'июля,': '07',
    'августа,': '08',
    'сентября,': '09',
    'октября,': '10',
    'ноября,': '11',
    'декабря,': '12'
}
days_of_week = {
    'Понедельник': 'Пн',
    'Вторник': 'Вт',
    'Среда': 'Ср',
    'Четверг': 'Чт',
    'Пятница': 'Пт',
    'Суббота': 'Сб',
    'Воскресенье': 'Вс'
}
rows_games = 2
games_dict = {}
# Цикл для перебора страниц с 1 по n
for page_num in range(1):

    url = base_url
    
    # Отправляем GET-запрос к странице
    response = requests.get(url)
    
    if response.status_code == 200:
        # Инициализируем объект BeautifulSoup для парсинга HTML
        soup = BeautifulSoup(response.text, 'html.parser')
    
        # Находим все блоки с датой игр
        date_of_game_divs = soup.find_all('div', class_='h3 h3-white h3-mb10')
        bar_info_divs = soup.find_all('div', class_='schedule-block-info-bar')
        results_pages = soup.find_all('div', class_='schedule-block-top')
        results_pages.reverse()
        date_of_game_divs.reverse()
        bar_info_divs.reverse()
        for j in range(len(results_pages)):
            link_to_game = results_pages[j].find('a', href=lambda href: href.startswith('/game-page?id='))
            day, month, day_of_week  = date_of_game_divs[j].get_text(strip=True).split()
            bar_info = ''.join(bar_info_divs[j].find_all(text=True, recursive=False))
            formatted_date = f"{day.zfill(2)}.{months[month]}"
            formatted_day_of_week = days_of_week[day_of_week]
            if link_to_game:

                game_number_div = link_to_game.find('div', class_='h2 h2-game-card')
                game_name_div = link_to_game.find('div', class_='h2 h2-game-card h2-left')
                game_number = game_number_div.text.strip()
                game_name = game_name_div.text.strip()
                
                game_info = (game_name + ' ' + game_number)
                if game_info not in games_dict:
                    games_dict[game_info] = 1
                else:
                    games_dict[game_info] += 1
                if 'стрим' not in game_info:
                    results_page_url = 'https://spb.quizplease.ru' + link_to_game['href']
                    other_page_response = requests.get(results_page_url)
            
                    if other_page_response.status_code == 200:
                        soup = BeautifulSoup(other_page_response.text, 'html.parser')

                        table_rows = soup.find_all('tr')

                        for gang_bang_team in table_rows:
                            # Находим все теги <td> в текущем теге <tr>
                            td_cells = gang_bang_team.find_all('td', class_='game-table__cell text')
                            td_cells_img = gang_bang_team.find_all('img', class_='game-table-rang')
                            if len(td_cells) >= 3:
                                team_placement = td_cells[0].text.strip()
                                if len(td_cells_img) >= 1:
                                    team_name = td_cells[1].text.strip()
                                else:
                                    team_name = td_cells[2].text.strip()
                                summary_points = td_cells[-1].text.strip()
                                ws.cell(row=rows_games, column=1, value=game_info)
                                ws.cell(row=rows_games, column=2, value=(bar_info.strip()))
                                ws.cell(row=rows_games, column=3, value=(formatted_date))
                                ws.cell(row=rows_games, column=4, value=(formatted_day_of_week))
                                ws.cell(row=rows_games, column=5, value=games_dict[game_info])
                                ws.cell(row=rows_games, column=6, value=team_name)
                                if summary_points != '':
                                    ws.cell(row=rows_games, column=7, value=float(summary_points))
                                else:
                                    ws.cell(row=rows_games, column=7, value=None)
                                ws.cell(row=rows_games, column=8, value=int(team_placement))
                                rows_games += 1
            
                    else:
                        print('Ошибка при получении другой страницы:', other_page_response.status_code)
            else:
                print('Ссылка на другую страницу не найдена.')
    else:
        print('Ошибка при получении страницы:', response.status_code)

wb.save("finals.xlsx")
wb.close()