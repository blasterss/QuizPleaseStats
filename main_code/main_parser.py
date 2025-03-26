import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

class GameDataScraper:
    def __init__(self, category, output_filename, input_filename):
        self.category = category
        self.output_filename = output_filename
        self.input_filename = input_filename
        self.months = {
            'января,': '01',
            'февраля,': '02',
            'марта,': '03',
            'апреля,': '04',
            'мая,': '05',
            'июня,': '06',
            'июля,': '07',
            'августа,': '08',
            'сентября,': '09',
            'октября,': '10',
            'ноября,': '11',
            'декабря,': '12'
        }
        self.days_of_week = {
            'Понедельник': 'Пн',
            'Вторник': 'Вт',
            'Среда': 'Ср',
            'Четверг': 'Чт',
            'Пятница': 'Пт',
            'Суббота': 'Сб',
            'Воскресенье': 'Вс'
        }
        self.games_dict = {}
        self.rows_games = 2
        self.kim_thematic_url = 'https://spb.quizplease.ru/schedule-past?QpGameSearch%5BcityId%5D=17&QpGameSearch%5Bmonth%5D=0&QpGameSearch%5Btype%5D=9&QpGameSearch%5Bbars%5D=all&page={}&per-page=12'
        self.kim_url = 'https://spb.quizplease.ru/schedule-past?QpGameSearch%5BcityId%5D=17&QpGameSearch%5Bmonth%5D=0&QpGameSearch%5Btype%5D=5&QpGameSearch%5Bbars%5D=all&page={}&per-page=12'
        self.classic_url = 'https://spb.quizplease.ru/schedule-past?QpGameSearch%5BcityId%5D=17&QpGameSearch%5Bmonth%5D=0&QpGameSearch%5Btype%5D=1&QpGameSearch%5Bbars%5D=all&page={}&per-page=12'
        self.classic_thematic_url = 'https://spb.quizplease.ru/schedule-past?QpGameSearch%5BcityId%5D=17&QpGameSearch%5Bmonth%5D=0&QpGameSearch%5Btype%5D=2&QpGameSearch%5Bbars%5D=all&page={}&per-page=12'

    def scrape_data(self, url, category):
        wb = load_workbook(self.input_filename)
        ws = wb.active
        self.rows_games = self.get_first_empty_row(ws)
        print(self.rows_games)
        current_game_ids = self.read_game_ids(category)

        #как-то контролировать количество страниц для ускорения работы приложения
        if category == 'kim':
            url = self.kim_url
            pages = 8 #первые игры были без оскаров
        elif category == 'thematic_kim':
            url = self.kim_thematic_url
            pages = 12 #первые игры относились к классике
        elif category == 'classic':
            url = self.classic_url
            pages = 30 #это реально дофига
        elif category == 'thematic_classic':
            url = self.classic_thematic_url
            pages = 9

        for page_num in range(pages, 0, -1):

            current_url = url.format(page_num)
            response = requests.get(current_url)

            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')
                date_of_game_divs, bar_info_divs, results_pages = self.extract_elements(soup)

                for j, result_page in enumerate(results_pages):
                    link_to_game = result_page.find('a', href=lambda href: href.startswith('/game-page?id='))
                    day, month, day_of_week = date_of_game_divs[j].get_text(strip=True).split()
                    bar_info = ''.join(bar_info_divs[j].find_all(text=True, recursive=False))
                    formatted_date = f"{day.zfill(2)}.{self.months[month]}.2024"
                    formatted_day_of_week = self.days_of_week[day_of_week]

                    if link_to_game:
                        game_info = self.extract_game_info(link_to_game)

                        if game_info not in self.games_dict:
                            self.games_dict[game_info] = 1
                        else:
                            self.games_dict[game_info] += 1

                        if ('стрим' not in game_info) and (link_to_game['href'] not in current_game_ids):
                            self.scrape_game_data(link_to_game, bar_info, formatted_date, formatted_day_of_week, game_info, ws, category)
                    else:
                        print('Ссылка на другую страницу не найдена.')
            else:
                print('Ошибка при получении страницы:', response.status_code)

        wb.save(self.output_filename)
        wb.close()

    def read_game_ids(self, category):
        path = self.get_game_id_path(category)
        with open(path, 'r') as file_txt:
            arr = [line.strip() for line in file_txt]
        return arr

    def get_game_id_path(self, category):
        if category == 'kim':
            path = 'games_ids/games_id_kim.txt'
        elif category == 'thematic_kim':
            path = 'games_ids/games_id_thematic_kim.txt'
        elif category == 'classic':
            path = 'games_ids/games_id_classic.txt'
        elif category == 'thematic_classic':
            path = 'games_ids/games_id_thematic_classic.txt'
        else:
            path = None
        return path
    
    def extract_elements(self, soup):
        date_of_game_divs = soup.find_all('div', class_='h3 h3-white h3-mb10')
        bar_info_divs = soup.find_all('div', class_='schedule-block-info-bar')
        results_pages = soup.find_all('div', class_='schedule-block-top')
        results_pages.reverse()
        date_of_game_divs.reverse()
        bar_info_divs.reverse()
        return date_of_game_divs, bar_info_divs, results_pages

    def extract_game_info(self, link_to_game):
        game_number_div = link_to_game.find('div', class_='h2 h2-game-card')
        game_name_div = link_to_game.find('div', class_='h2 h2-game-card h2-left')
        game_number = game_number_div.text.strip()
        game_name = game_name_div.text.strip()
        return f"{game_name} {game_number}"

    def get_first_empty_row(self, sheet):
        max_row = sheet.max_row
        for i in range(1, max_row + 1):
            if sheet.cell(row=i, column=1).value is None:
                return i
        return max_row + 1

    #Попробовать внедрить данную функцию и есть ли смысл
    def filling_data(self, team_placement, team_name, final_round, summary_points, bar_info, formatted_date, formatted_day_of_week, game_info, ws):
        ws.cell(row=self.rows_games, column=1, value=game_info)
        ws.cell(row=self.rows_games, column=2, value=(bar_info.strip()))
        ws.cell(row=self.rows_games, column=3, value=(formatted_date))
        ws.cell(row=self.rows_games, column=4, value=(formatted_day_of_week))
        ws.cell(row=self.rows_games, column=5, value=self.games_dict[game_info])
        ws.cell(row=self.rows_games, column=6, value=team_name)
        if summary_points != '':
            ws.cell(row=self.rows_games, column=7, value=float(summary_points))
        else:
            ws.cell(row=self.rows_games, column=7, value=None)
        if final_round != '':
            ws.cell(row=self.rows_games, column=8, value=float(final_round))
        else:
            ws.cell(row=self.rows_games, column=8, value=None)
        ws.cell(row=self.rows_games, column=9, value=int(team_placement))
        self.rows_games += 1

    def scrape_game_data(self, link_to_game, bar_info, formatted_date, formatted_day_of_week, game_info, ws, category):
        results_page_url = 'https://spb.quizplease.ru' + link_to_game['href']
        other_page_response = requests.get(results_page_url)

        if other_page_response.status_code == 200:
            soup = BeautifulSoup(other_page_response.text, 'html.parser')

            path = self.get_game_id_path(category)
            with open(path, 'a') as file_txt:
                file_txt.write(link_to_game['href'] + '\n')

            table_rows = soup.find_all('tr')
            #td_head_row = team_data.find_all('td', class_='label game-table__header-cell')
            #team_index = td_head_row.index('1 раунд')

            for team_data in table_rows:

                td_cells = team_data.find_all('td', class_='game-table__cell text')
                td_cells_img = team_data.find_all('img', class_='game-table-rang')
                count = 1

                if (category == 'kim') or (category == 'thematic_kim'):

                    if len(td_cells) >= 4:
                        team_placement = td_cells[0].text.strip()
                        if len(td_cells_img) == 2:
                            team_name = td_cells[1].text.strip()
                            count = 2
                        elif len(td_cells_img) == 1:
                            team_name = td_cells[2].text.strip()
                            count = 3
                        else:
                            team_name = td_cells[3].text.strip()
                            count = 4
                        
                        #Добавлять статистику по раундам

                        final_round = td_cells[-2].text.strip()                            
                        summary_points = td_cells[-1].text.strip()
                        summary_points = summary_points.replace(',', '.')
                        ws.cell(row=self.rows_games, column=1, value=game_info)
                        ws.cell(row=self.rows_games, column=2, value=(bar_info.strip()))
                        ws.cell(row=self.rows_games, column=3, value=(formatted_date))
                        ws.cell(row=self.rows_games, column=4, value=(formatted_day_of_week))
                        ws.cell(row=self.rows_games, column=5, value=self.games_dict[game_info])
                        ws.cell(row=self.rows_games, column=6, value=team_name)
                        if final_round != '':
                            ws.cell(row=self.rows_games, column=7, value=float(final_round))
                        else:
                            ws.cell(row=self.rows_games, column=7, value=None)
                        if summary_points != '' and summary_points.find(','):
                            ws.cell(row=self.rows_games, column=8, value=float(summary_points))
                        else:
                            ws.cell(row=self.rows_games, column=8, value=None)
                        ws.cell(row=self.rows_games, column=9, value=int(team_placement))
                        self.rows_games += 1
                else:
                     if len(td_cells) >= 3:
                        team_placement = td_cells[0].text.strip()
                        if len(td_cells_img) >= 1:
                            team_name = td_cells[1].text.strip()
                        else:
                            team_name = td_cells[2].text.strip()

                        final_round = td_cells[-2].text.strip() 
                        summary_points = td_cells[-1].text.strip()
                        ws.cell(row=self.rows_games, column=1, value=game_info)
                        ws.cell(row=self.rows_games, column=2, value=(bar_info.strip()))
                        ws.cell(row=self.rows_games, column=3, value=(formatted_date))
                        ws.cell(row=self.rows_games, column=4, value=(formatted_day_of_week))
                        ws.cell(row=self.rows_games, column=5, value=self.games_dict[game_info])
                        ws.cell(row=self.rows_games, column=6, value=team_name)
                        if final_round != '':
                            ws.cell(row=self.rows_games, column=7, value=float(final_round))
                        else:
                            ws.cell(row=self.rows_games, column=7, value=None)
                        if summary_points != '':
                            ws.cell(row=self.rows_games, column=8, value=float(summary_points))
                        else:
                            ws.cell(row=self.rows_games, column=8, value=None)
                        ws.cell(row=self.rows_games, column=9, value=int(team_placement))
                        self.rows_games += 1
        else:
            print('Ошибка при получении другой страницы:', other_page_response.status_code)


